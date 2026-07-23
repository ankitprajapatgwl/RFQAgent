"""Read stored email attachments back into content the extractor can analyse.

Attachments can be *anything* a supplier sends, and the agent must read them all
(Requirement: "API MUST read all files automatically and extract details"). Each
attachment is turned into the richest representation the model can consume:

* **Text-shaped files** (plain text, CSV, JSON, HTML, Markdown, XML, ...) are
  decoded from disk and rendered inline as text.
* **Spreadsheets** (``.xlsx``/``.xlsm``) are parsed with ``openpyxl`` and their
  cells rendered as delimited text rows.
* **PDFs and images** are handed to Claude natively as ``document`` / ``image``
  content blocks (base64) — Claude reads them directly, including scanned pages,
  screenshots, tables, and charts, without any extra OCR/PDF-parsing dependency.
* **Zip archives** are expanded in-memory and every member is processed the same
  way (one level deep), so a zipped quote-pack is read too.
* Anything still undecodable (unknown binary types) is represented by a short,
  honest placeholder noting the filename and type rather than being dropped.

All limits are bounded — per-file and total for text, and per-file/total bytes
plus a file count for native media — so a large or malicious attachment can never
blow the LLM token budget or the Anthropic request-size limit (coding standards:
token/cost guardrails; zip-bomb guards).
"""

from __future__ import annotations

import base64
import io
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from src.config import Settings
from src.observability import get_logger

logger = get_logger(__name__)

# ── Bounds that keep attachment content from blowing the LLM token budget ──────
# Text (decoded inline).
_MAX_CHARS_PER_FILE = 12_000
_MAX_TOTAL_CHARS = 40_000

# Native media (PDFs + images sent to the model as document/image blocks). The
# Anthropic request cap is 32 MB total; we stay well under it and cap the count
# so a burst of large attachments cannot explode the request size or token cost.
_MAX_MEDIA_BYTES_PER_FILE = 4_500_000
_MAX_TOTAL_MEDIA_BYTES = 24_000_000
_MAX_MEDIA_FILES = 20

# Absolute cap on how much of any single file we ever read into memory.
_MAX_READ_BYTES = 32_000_000

# Spreadsheet rendering bounds (per sheet).
_MAX_SHEET_ROWS = 200
_MAX_SHEET_COLS = 50

# Zip expansion guards (defence against zip bombs).
_MAX_ZIP_MEMBERS = 50
_MAX_ZIP_TOTAL_UNCOMPRESSED_BYTES = 60_000_000
_MAX_ZIP_DEPTH = 1

# Content types (and filename suffixes) we treat as safely text-decodable.
_TEXT_CONTENT_TYPES = {
    "application/json",
    "application/xml",
    "application/csv",
    "application/x-ndjson",
    "application/yaml",
    "application/x-yaml",
}
_TEXT_SUFFIXES = (
    ".txt",
    ".text",
    ".md",
    ".markdown",
    ".csv",
    ".tsv",
    ".json",
    ".xml",
    ".html",
    ".htm",
    ".yaml",
    ".yml",
    ".log",
    ".rtf",
)

# Spreadsheet formats openpyxl can parse (legacy ``.xls`` is a different binary
# format openpyxl cannot read — it falls through to a placeholder).
_SPREADSHEET_SUFFIXES = (".xlsx", ".xlsm")

# Image formats Claude accepts as native ``image`` blocks, mapped to media type.
_IMAGE_MEDIA_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".gif": "image/gif",
    ".webp": "image/webp",
}


@dataclass(frozen=True)
class AttachmentRef:
    """The minimum an attachment needs to be located on disk and labelled.

    Attributes:
        filename: Original filename (for labelling in the prompt).
        url: The served URL stored on the attachment row; its basename is the
            on-disk filename under :attr:`Settings.attachments_dir`.
        content_type: MIME type, when known.
        size_bytes: Stored size, when known.
    """

    filename: str
    url: str
    content_type: str | None = None
    size_bytes: int | None = None


@dataclass
class AttachmentContent:
    """Everything read out of an email's attachments, ready for the prompt.

    Attributes:
        text: A human-readable, clearly-delimited block describing every
            attachment — inline for text/spreadsheets, a short marker for
            PDFs/images (whose bytes travel in :attr:`media_blocks`), and a
            placeholder for anything undecodable. Empty when there are none.
        media_blocks: Anthropic content blocks (``document``/``image``) for
            attachments read natively by the model, in the same order they are
            referenced in :attr:`text`.
    """

    text: str = ""
    media_blocks: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class _Accumulator:
    """Running totals used to enforce the bounds across all attachments."""

    parts: list[str] = field(default_factory=list)
    media_blocks: list[dict[str, Any]] = field(default_factory=list)
    text_chars: int = 0
    media_bytes: int = 0

    def add_text(self, header: str, body: str) -> None:
        """Append a labelled text section, clipped to the remaining budget."""
        remaining = _MAX_TOTAL_CHARS - self.text_chars
        if remaining <= 0:
            self.parts.append(f"{header} ---\n[skipped — attachment text budget reached]")
            return
        clipped = body[: min(_MAX_CHARS_PER_FILE, remaining)]
        if len(clipped) < len(body):
            clipped += "\n[... truncated ...]"
        self.text_chars += len(clipped)
        self.parts.append(f"{header} ---\n{clipped}")

    def add_note(self, header: str, note: str) -> None:
        """Append a labelled marker/placeholder line (does not count toward text)."""
        self.parts.append(f"{header} ---\n{note}")

    def add_media(self, header: str, block: dict[str, Any], kind: str, num_bytes: int) -> bool:
        """Append a native media block if it fits the byte/count budgets.

        Returns:
            ``True`` if the block was added; ``False`` if a budget was hit (a
            placeholder note is written either way).
        """
        if len(self.media_blocks) >= _MAX_MEDIA_FILES:
            self.add_note(header, f"[{kind} skipped — attachment file limit reached]")
            return False
        if num_bytes > _MAX_MEDIA_BYTES_PER_FILE:
            self.add_note(header, f"[{kind} skipped — too large to read ({num_bytes} bytes)]")
            return False
        if self.media_bytes + num_bytes > _MAX_TOTAL_MEDIA_BYTES:
            self.add_note(header, f"[{kind} skipped — attachment size budget reached]")
            return False
        self.media_bytes += num_bytes
        self.media_blocks.append(block)
        self.add_note(header, f"[{kind} provided as a separate content block for you to read]")
        return True


def _looks_textual(filename: str, content_type: str | None) -> bool:
    """Return whether a file is likely safe to decode as UTF-8 text."""
    ct = (content_type or "").lower()
    if ct.startswith("text/") or ct in _TEXT_CONTENT_TYPES:
        return True
    return filename.lower().endswith(_TEXT_SUFFIXES)


def _image_media_type(filename: str, content_type: str | None) -> str | None:
    """Return the Anthropic image media type for a file, or ``None`` if unsupported."""
    ct = (content_type or "").lower()
    if ct in _IMAGE_MEDIA_TYPES.values():
        return ct
    suffix = Path(filename.lower()).suffix
    return _IMAGE_MEDIA_TYPES.get(suffix)


def _is_pdf(filename: str, content_type: str | None) -> bool:
    """Return whether a file is a PDF (by content type or extension)."""
    ct = (content_type or "").lower()
    return ct == "application/pdf" or filename.lower().endswith(".pdf")


def _is_spreadsheet(filename: str, content_type: str | None) -> bool:
    """Return whether a file is an openpyxl-readable spreadsheet."""
    if filename.lower().endswith(_SPREADSHEET_SUFFIXES):
        return True
    ct = (content_type or "").lower()
    return "spreadsheetml" in ct  # e.g. application/vnd.openxmlformats-...spreadsheetml.sheet


def _is_zip(filename: str, content_type: str | None) -> bool:
    """Return whether a file is a zip archive (by content type or extension)."""
    ct = (content_type or "").lower()
    return ct in {"application/zip", "application/x-zip-compressed"} or filename.lower().endswith(
        ".zip"
    )


def _render_spreadsheet(data: bytes) -> str:
    """Render a spreadsheet's cells as delimited text rows (bounded per sheet)."""
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sheets: list[str] = []
        for sheet in workbook.worksheets:
            lines = [f"# Sheet: {sheet.title}"]
            for r, row in enumerate(sheet.iter_rows(values_only=True)):
                if r >= _MAX_SHEET_ROWS:
                    lines.append("[... more rows truncated ...]")
                    break
                cells = ["" if v is None else str(v) for v in row[:_MAX_SHEET_COLS]]
                lines.append(" | ".join(cells))
            sheets.append("\n".join(lines))
        return "\n\n".join(sheets)
    finally:
        workbook.close()


def _process(
    *,
    acc: _Accumulator,
    header: str,
    filename: str,
    content_type: str | None,
    data: bytes,
    depth: int,
) -> None:
    """Classify one attachment's bytes and add its content to ``acc``.

    Shared by on-disk attachments and zip members so both are read identically.
    """
    if _looks_textual(filename, content_type):
        acc.add_text(header, data.decode("utf-8", errors="replace"))
        return

    if _is_spreadsheet(filename, content_type):
        try:
            acc.add_text(header, _render_spreadsheet(data))
        except (InvalidFileException, zipfile.BadZipFile, KeyError, ValueError) as exc:
            logger.warning("Could not parse spreadsheet %s: %s", filename, exc)
            acc.add_note(header, "[could not read spreadsheet]")
        return

    if _is_pdf(filename, content_type):
        block = {
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": base64.standard_b64encode(data).decode("ascii"),
            },
        }
        acc.add_media(header, block, "PDF document", len(data))
        return

    image_media_type = _image_media_type(filename, content_type)
    if image_media_type is not None:
        block = {
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": image_media_type,
                "data": base64.standard_b64encode(data).decode("ascii"),
            },
        }
        acc.add_media(header, block, "image", len(data))
        return

    if _is_zip(filename, content_type):
        _process_zip(acc=acc, header=header, filename=filename, data=data, depth=depth)
        return

    acc.add_note(header, "[binary attachment — not a readable type]")


def _process_zip(*, acc: _Accumulator, header: str, filename: str, data: bytes, depth: int) -> None:
    """Expand a zip archive in memory and process each member (one level deep)."""
    if depth >= _MAX_ZIP_DEPTH:
        acc.add_note(header, "[nested zip archive — not expanded]")
        return
    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile:
        acc.add_note(header, "[could not read zip archive]")
        return

    with archive:
        members = [info for info in archive.infolist() if not info.is_dir()]
        acc.add_note(header, f"[zip archive with {len(members)} file(s) — expanded below]")
        uncompressed = 0
        for info in members[:_MAX_ZIP_MEMBERS]:
            uncompressed += info.file_size
            if uncompressed > _MAX_ZIP_TOTAL_UNCOMPRESSED_BYTES:
                acc.add_note(header, "[remaining zip members skipped — archive too large]")
                break
            member_header = f"--- {filename} → {info.filename}"
            try:
                member_data = archive.read(info)
            except (zipfile.BadZipFile, RuntimeError, OSError) as exc:
                logger.warning("Could not read zip member %s: %s", info.filename, exc)
                acc.add_note(member_header, "[could not read zip member]")
                continue
            _process(
                acc=acc,
                header=member_header,
                filename=info.filename,
                content_type=None,
                data=member_data,
                depth=depth + 1,
            )
        if len(members) > _MAX_ZIP_MEMBERS:
            acc.add_note(header, f"[{len(members) - _MAX_ZIP_MEMBERS} more zip member(s) skipped]")


def read_attachments(settings: Settings, attachments: list[AttachmentRef]) -> AttachmentContent:
    """Read every attachment into text + native media blocks for the prompt.

    Text-shaped files and spreadsheets are rendered inline; PDFs and images are
    returned as native Anthropic ``document``/``image`` blocks; zips are expanded
    and their members read the same way. Every failure (missing file, unreadable
    bytes, bad archive) is logged and noted rather than aborting the batch.

    Args:
        settings: Application settings (locates the attachments directory).
        attachments: The attachments to read.

    Returns:
        An :class:`AttachmentContent` with the delimited text block and the
        ordered list of media blocks (both empty when there are no attachments).
    """
    if not attachments:
        return AttachmentContent()

    directory = settings.attachments_dir
    acc = _Accumulator()

    for index, ref in enumerate(attachments, start=1):
        header = f"--- Attachment {index}: {ref.filename} ({ref.content_type or 'unknown type'})"

        name = (ref.url or "").rsplit("/", 1)[-1]
        path = directory / name if name else None
        if path is None or not path.exists():
            acc.add_note(header, "[attachment file not found on disk]")
            continue

        try:
            if path.stat().st_size > _MAX_READ_BYTES:
                acc.add_note(header, "[attachment skipped — file too large to read]")
                continue
            data = path.read_bytes()
        except OSError as exc:
            logger.warning("Could not read attachment %s: %s", name, exc)
            acc.add_note(header, "[could not read attachment]")
            continue

        _process(
            acc=acc,
            header=header,
            filename=ref.filename,
            content_type=ref.content_type,
            data=data,
            depth=0,
        )

    return AttachmentContent(text="\n\n".join(acc.parts), media_blocks=acc.media_blocks)
